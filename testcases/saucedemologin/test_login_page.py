import openpyxl
import pytest

from pages.saucedemo.sauce_demo_login_page import LoginPage
from testcases.test_base_class import TestBaseClass


def get_test_data_set(sheet_name):
    '''
    To get data from excel file
    :param sheet_name:
    :return: array of dictionaries
    '''
    lis = []
    book = openpyxl.load_workbook("./library/TestData.xlsx")
    sheet = book[sheet_name]

    data = []
    for row in range(2, sheet.max_row):
        if sheet.cell(row, 1).value is None:
            break
        elm = {}
        for col in range(1, sheet.max_column + 1):
            elm[sheet.cell(row=1, column=col).value] = sheet.cell(row, col).value
        data.append(elm)
    book.close()
    return data


testdata = get_test_data_set("saucedemo")


@pytest.mark.parametrize('testdata', testdata, scope="class")
class TestErrorLogin(TestBaseClass):

    def test_login_page_error_validations(self,testdata):
        login_page= LoginPage(self.driver)
        login_page.login_to_swag_labs(testdata["username"], testdata["password"])




