import openpyxl
import pytest

from pages.saucedemo.sauce_demo_login_page import LoginPage
from testcases.test_base_class import TestBaseClass


def get_test_data_set(sheet_name):
    '''
    To get data from excel file
    :param sheet_name:
    :return: array of dictionaries
    '''
    lis = []
    book = openpyxl.load_workbook("./library/TestData.xlsx")
    sheet = book[sheet_name]

    data = []
    for row in range(2, sheet.max_row):
        if sheet.cell(row, 1).value is None:
            break
        elm = {}
        for col in range(1, sheet.max_column + 1):
            elm[sheet.cell(row=1, column=col).value] = sheet.cell(row, col).value
        data.append(elm)
    book.close()
    return data


testdata = get_test_data_set("saucedemo")


@pytest.mark.parametrize("testdata", testdata, scope="class")
class TestSuccessLogin(TestBaseClass):

    @pytest.mark.usefixtures("logout")
    def test_standard_user_login(self, testdata):
        global sauce_demo_login_page
        sauce_demo_login_page = LoginPage(self.driver)
        sauce_demo_login_page.login_to_swag_labs(testdata["standard_username"], testdata["password"])

    def test_locked_user_login(self, testdata):
        sauce_demo_login_page = LoginPage(self.driver)
        sauce_demo_login_page.login_to_swag_labs(testdata["locked_username"], testdata["password"])
        assert "Epic sadface: Sorry, this user has been locked out." == sauce_demo_login_page.get_error_message()



    @pytest.fixture(scope="function")
    def logout(self):

        yield
        # import pdb
        # pdb.set_trace()
        sauce_demo_login_page = LoginPage(self.driver)
        sauce_demo_login_page.logout()