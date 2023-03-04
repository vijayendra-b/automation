import re
import allure
import openpyxl
import pytest
from allure_commons.types import AttachmentType


@pytest.mark.usefixtures("setup")
class BaseClass:

    def teardown_method(self):
        """
        to capture screenshot
        :return:
        """
        allure.attach(self.driver.get_screenshot_as_png(), "Failed Test", attachment_type=AttachmentType.PNG)
        # attach(data=self.driver.get_screenshot_as_png())

    def get_test_data(self, sheet_name):
        """
        Finds and returns the element specified by locator
        :param sheet_name: the test sheet name in the TestData.xlsx file
        :return: element
        """
        dic = {}
        book = openpyxl.load_workbook("./library/TestData.xlsx")
        sheet = book[sheet_name]
        for j in range(1, sheet.max_column):
            if sheet.cell(row=1, column=j).value is None:
                break
            dic[sheet.cell(row=1, column=j).value] = sheet.cell(row=2, column=j).value
        book.close()
        return dic

    def get_test_data_set(self, sheet_name):
        '''
        To get data from excel file
        :param sheet_name:
        :return: array of dictionaries
        '''
        lis = []
        book = openpyxl.load_workbook("./library/TestData.xlsx")
        sheet = book[sheet_name]

        data = []
        for row in range(2, sheet.max_row):
            if sheet.cell(row, 1).value is None:
                break
            elm = {}
            for col in range(1, sheet.max_column + 1):
                elm[sheet.cell(row=1, column=col).value] = sheet.cell(row, col).value
            data.append(elm)
        book.close()
        return data

    def is_year_in_text(self, text):
        match = re.search('\\d{4}', text)
        return match is not None
